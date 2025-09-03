from openpyxl import load_workbook

file_path = "your_file.xlsx"
sheet_name = "Other expense"
output_path = "formatted_output.xlsx"

# Load workbook and target sheet
wb = load_workbook(file_path)
ws = wb[sheet_name]

# Helper: safe conversion to float
def to_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None

# Step 1: Prefix subcategories in column A (and text cells of subcategories)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
    cell = row[0]
    if cell.value is None:
        continue
    outline_level = ws.row_dimensions[cell.row].outlineLevel
    if outline_level > 0:
        if not str(cell.value).startswith("-"):
            cell.value = "-" * outline_level + " " + str(cell.value)

# Step 2: Aggregate children into parent across ALL numeric columns (B → end)
for row in range(1, ws.max_row + 1):
    outline_level = ws.row_dimensions[row].outlineLevel

    if outline_level == 0:  # parent row
        # Find contiguous children
        child_row = row + 1
        while child_row <= ws.max_row and ws.row_dimensions[child_row].outlineLevel > 0:
            child_row += 1

        if child_row > row + 1:  # if children exist
            for col in range(2, ws.max_column + 1):  # B onwards
                parent_val = to_float(ws.cell(row=row, column=col).value)
                child_sum = 0.0
                has_child_number = False

                # sum children only (col by col)
                for r in range(row + 1, child_row):
                    val = to_float(ws.cell(row=r, column=col).value)
                    if val is not None:
                        child_sum += val
                        has_child_number = True

                if has_child_number:
                    if parent_val is None:
                        # parent empty → replace with sum
                        ws.cell(row=row, column=col).value = child_sum
                    else:
                        # replace only if child sum > parent
                        if parent_val < child_sum:
                            ws.cell(row=row, column=col).value = child_sum
                        else:
                            ws.cell(row=row, column=col).value = parent_val

# Save updated workbook
wb.save(output_path)

print(f"✅ Prefixing + Column-wise aggregation complete! Updated file saved as: {output_path}")